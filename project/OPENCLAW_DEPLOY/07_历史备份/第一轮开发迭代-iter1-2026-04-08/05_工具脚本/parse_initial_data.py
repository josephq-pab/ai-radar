from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

BASE = Path('/home/admin/.openclaw/workspace-ai-radar/project/OPENCLAW_DEPLOY')
RAW = BASE / 'data' / 'raw'
PROCESSED = BASE / 'data' / 'processed'

BANK_MAP = {
    '平安': '平安银行',
    '浦发': '浦发银行',
    '民生': '民生银行',
    '华夏': '华夏银行',
    '光大': '光大银行',
    '兴业': '兴业银行',
    '招行': '招商银行',
    '招商': '招商银行',
    '中信': '中信银行',
    '交行': '交通银行',
    '交通': '交通银行',
}

FOCUS_BANKS = ['平安银行', '招商银行', '兴业银行', '中信银行', '浦发银行', '民生银行', '华夏银行', '光大银行']


def norm_bank(name: Any) -> str:
    text = str(name or '').strip()
    return BANK_MAP.get(text, text)


def to_num(value: Any) -> Any:
    if value is None or value == '':
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(',', '')
    try:
        return float(text)
    except Exception:
        return text


def excel_serial_to_date(serial: int) -> str:
    """将 Excel 日期序列号转换为 YYYY-MM-DD 格式"""
    from datetime import date, timedelta
    return (date(1899, 12, 30) + timedelta(days=int(serial))).strftime('%Y-%m-%d')


def parse_date_token(value: Any) -> str | None:
    text = str(value or '').strip()
    # 格式1: YYYY-MM-DD 或 YYYY/MM/DD 或 YYYYMMDD
    m = re.search(r'(20\d{2})[/-]?(\d{1,2})[/-]?(\d{1,2})', text)
    if m:
        y, mo, d = m.groups()
        return f'{int(y):04d}-{int(mo):02d}-{int(d):02d}'
    # 格式2: "3月31日"（新文件标题行中的日期）
    m = re.search(r'(\d{1,2})月(\d{1,2})日', text)
    if m:
        mo, d = m.groups()
        return f'2026-{int(mo):02d}-{int(d):02d}'
    # 格式3: "2026年3月31日"
    m = re.search(r'(20\d{2})年(\d{1,2})月(\d{1,2})日', text)
    if m:
        y, mo, d = m.groups()
        return f'{int(y):04d}-{int(mo):02d}-{int(d):02d}'
    return None


def pct_text(value: Any) -> str:
    if value is None or value == '':
        return '--'
    try:
        return f'{float(value) * 100:.2f}%'
    except Exception:
        return str(value)


def num_text(value: Any, digits: int = 2) -> str:
    if value is None or value == '':
        return '--'
    try:
        return f'{float(value):,.{digits}f}'
    except Exception:
        return str(value)


def write_json(name: str, payload: Any) -> None:
    path = PROCESSED / name
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')


def write_csv(name: str, rows: list[dict[str, Any]]) -> None:
    path = PROCESSED / name
    if not rows:
        path.write_text('', encoding='utf-8')
        return
    headers = sorted({k for row in rows for k in row.keys()})
    with path.open('w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def parse_deposit() -> list[dict[str, Any]]:
    # 优先匹配 2026 年 3 月新文件（存款对标数据_2026年3月.xlsx）
    # 模式：含"存款"+"对标"，按文件名倒序取最新月份
    # TODO(临时方案)：依赖文件名倒序不够稳健，后续改为显式解析文件名年月，
    #                 或在元数据中存储文件归属月份，避免文件名格式变化导致逻辑失效。
    candidates = sorted(
        RAW.glob('*存款*对标*.xlsx'),
        key=lambda f: f.name,
        reverse=True
    )
    if not candidates:
        raise FileNotFoundError('找不到存款对标数据文件')
    file = candidates[0]
    wb = load_workbook(file, data_only=True, read_only=True)
    # 动态识别 sheet：优先找含"存款"的 sheet，否则用第一个
    sheet_names = wb.sheetnames
    ws_name = next((s for s in sheet_names if '存款' in s), sheet_names[0])
    ws = wb[ws_name]
    rows = list(ws.iter_rows(values_only=True))
    date_label = str(rows[0][0] or '')
    observed_at = None
    m = re.search(r'(\d+)月(\d+)日', date_label)
    if m:
        observed_at = f'2026-{int(m.group(1)):02d}-{int(m.group(2)):02d}'

    metric_groups = [
        ('境内本外币存款', 'balance_total'),
        ('境内人民币存款', 'balance_rmb'),
        ('境内外币存款', 'balance_fx'),
    ]
    out = []
    for row in rows[4:]:
        bank = norm_bank(row[0])
        if not bank:
            continue
        values = list(row)
        for idx, (metric_name, metric_code) in enumerate(metric_groups):
            base = 1 + idx * 4
            out.append({
                'dataset': 'deposit_benchmark',
                'observedAt': observed_at,
                'bank': bank,
                'metricGroup': metric_name,
                'metricCode': metric_code,
                'currentValue': to_num(values[base] if base < len(values) else None),
                'monthChange': to_num(values[base + 1] if base + 1 < len(values) else None),
                'yearChange': to_num(values[base + 2] if base + 2 < len(values) else None),
                'yearGrowth': to_num(values[base + 3] if base + 3 < len(values) else None),
                'sourceFile': file.name,
                'sourceSheet': '余额',
            })
    return out


def parse_loan() -> list[dict[str, Any]]:
    # 优先匹配 2026 年 3 月新文件（贷款对标数据_2026年3月.xlsx）
    # 模式：含"贷款"+"对标"，按文件名倒序取最新月份
    # TODO(临时方案)：同上，应改为显式解析年月而非依赖文件名倒序。
    loan_files = sorted(
        RAW.glob('*贷款*对标*.xlsx'),
        key=lambda f: f.name,
        reverse=True
    )
    if not loan_files:
        raise FileNotFoundError('找不到贷款对标数据文件')
    file = loan_files[0]
    wb = load_workbook(file, data_only=True, read_only=True)
    out: list[dict[str, Any]] = []

    # 动态识别 sheet：优先找含"贷款"的 sheet，否则用第一个
    sheet_names = wb.sheetnames
    ws_name = next((s for s in sheet_names if '贷款' in s), sheet_names[0])
    ws = wb[ws_name]
    rows = list(ws.iter_rows(values_only=True))
    # 新文件：日期在 rows[0][0] 的标题里（格式："3月31日同业…"）
    # 旧文件：日期在 rows[1][1]，但可能是 YYYYMM 数字（202412 等）
    observed_at = None
    # 先试 rows[0][0]（新格式，如"3月31日"）
    if rows and rows[0]:
        observed_at = parse_date_token(rows[0][0])
    # 再试 rows[1][1]（旧格式，或 YYYYMM 数字列）
    if not observed_at and len(rows) > 1 and len(rows[1]) > 1:
        raw = rows[1][1]
        if isinstance(raw, (int, float)):
            # YYYYMM 格式数字：202412 → 2024-12
            s = str(int(raw))
            if len(s) == 6:
                observed_at = f'{s[:4]}-{s[4:6]}-01'
        else:
            observed_at = parse_date_token(raw)
    if not observed_at:
        observed_at = '2026-02-28'

    current_category = None
    for row in rows[2:]:
        vals = list(row)
        if len(vals) < 3:
            continue
        if vals[2] and not vals[1]:
            current_category = str(vals[2]).strip()
            continue
        bank = norm_bank(vals[1])
        if not bank:
            continue
        # 跳过表头行（norm 后 bank='银行'）
        if bank == '银行':
            continue
        # 跳过 currentValue 为非数值的行（防含中文说明文字的行）
        raw_cv = vals[2]
        if raw_cv is not None and isinstance(raw_cv, str):
            if not all(c in '0123456789.-+eE' for c in raw_cv.strip()):
                continue
        if current_category:
            out.append({
                'dataset': 'loan_benchmark',
                'observedAt': observed_at,
                'bank': bank,
                'metricGroup': current_category,
                'metricCode': 'loan_balance',
                'currentValue': to_num(vals[2]),
                'monthChange': to_num(vals[3]),
                'yearChange': to_num(vals[4]),
                'monthRank': to_num(vals[5]),
                'yearGrowth': to_num(vals[6]),
                'growthRank': to_num(vals[7]),
                'sourceFile': file.name,
                'sourceSheet': '表格展示',
            })
        if len(vals) >= 12 and any(v not in (None, '') for v in vals[8:12]):
            out.append({
                'dataset': 'loan_benchmark',
                'observedAt': observed_at,
                'bank': bank,
                'metricGroup': '一手福费廷',
                'metricCode': 'forfaiting_balance',
                'currentValue': to_num(vals[8]),
                'monthChange': to_num(vals[9]),
                'yearChange': to_num(vals[10]),
                'monthRank': to_num(vals[11]),
                'yearGrowth': None,
                'growthRank': None,
                'sourceFile': file.name,
                'sourceSheet': '表格展示',
            })
    return out


def parse_rates() -> list[dict[str, Any]]:
    file = next(RAW.glob('*贷款利率*.xlsx'))
    wb = load_workbook(file, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    for row in rows[1:]:
        vals = list(row)
        bank = norm_bank(vals[1] if len(vals) > 1 else None)
        if not bank:
            continue
        raw_observed = vals[0]
        # 转换 Excel 日期序列号或纯文本日期
        observed = parse_date_token(raw_observed)
        if observed is None and raw_observed is not None:
            # 尝试 Excel 序列号（纯数字）
            try:
                observed = excel_serial_to_date(int(raw_observed))
            except Exception:
                observed = None
        out.append({
            'dataset': 'loan_rate',
            'observedAt': observed,
            'bank': bank,
            'metricGroup': '贷款利率',
            'metricCode': 'weighted_avg_rate',
            'monthlyRate': to_num(vals[2] if len(vals) > 2 else None),
            'ytdRate': to_num(vals[3] if len(vals) > 3 else None),
            'sourceFile': file.name,
            'sourceSheet': ws.title,
        })
    return out


def dedup_records(records: list[dict[str, Any]], key_fields: list[str]) -> list[dict[str, Any]]:
    """按指定字段去重，保留每组第一条"""
    seen = set()
    result = []
    for r in records:
        key = tuple(r.get(f) for f in key_fields)
        if key not in seen:
            seen.add(key)
            result.append(r)
    return result


def build_summary(deposits, loans, rates):
    dep_total = [x for x in deposits if x['metricCode'] == 'balance_total' and x['bank'] in FOCUS_BANKS]
    dep_total_sorted = sorted(dep_total, key=lambda x: (x['currentValue'] or 0), reverse=True)
    loan_general = [x for x in loans if x['metricCode'] == 'loan_balance' and x['bank'] in FOCUS_BANKS]
    loan_general_sorted = sorted(loan_general, key=lambda x: (x['currentValue'] or 0), reverse=True)
    rate_sorted = sorted([x for x in rates if x['bank'] in FOCUS_BANKS], key=lambda x: (x['monthlyRate'] or 999))
    return {
        'focusBanks': FOCUS_BANKS,
        'depositTop': dep_total_sorted[:8],
        'loanTop': loan_general_sorted[:8],
        'rateRankingLowToHigh': rate_sorted,
    }


def build_quality_report(deposits, loans, rates):
    return {
        'depositRecords': len(deposits),
        'loanRecords': len(loans),
        'rateRecords': len(rates),
        'depositBanks': sorted({x['bank'] for x in deposits if x['bank']}),
        'loanBanks': sorted({x['bank'] for x in loans if x['bank']}),
        'rateBanks': sorted({x['bank'] for x in rates if x['bank']}),
        'missingObservedAtRate': [x['bank'] for x in rates if not x.get('observedAt')],
    }


def build_dashboard_payload(summary):
    deposit_top = summary['depositTop'][:8]
    loan_top = summary['loanTop'][:8]
    rate_low = summary['rateRankingLowToHigh'][:8]
    return {
        'kpis': {
            'focusBanks': len(summary['focusBanks']),
            'depositRecords': len(summary['depositTop']),
            'loanRecords': len(summary['loanTop']),
            'rateRecords': len(summary['rateRankingLowToHigh']),
        },
        'depositTop5': [
            {
                'bank': x['bank'],
                'metricGroup': x['metricGroup'],
                'currentValueText': num_text(x['currentValue']),
                'monthChangeText': num_text(x['monthChange']),
                'yearChangeText': num_text(x['yearChange']),
                'yearGrowthText': pct_text(x['yearGrowth']),
            } for x in deposit_top
        ],
        'loanTop5': [
            {
                'bank': x['bank'],
                'metricGroup': x['metricGroup'],
                'currentValueText': num_text(x['currentValue']),
                'monthChangeText': num_text(x['monthChange']),
                'yearChangeText': num_text(x['yearChange']),
                'yearGrowthText': pct_text(x['yearGrowth']),
            } for x in loan_top
        ],
        'rateLowToHigh5': [
            {
                'bank': x['bank'],
                'monthlyRateText': pct_text(x['monthlyRate'] / 100 if x['monthlyRate'] and x['monthlyRate'] > 1 else x['monthlyRate']),
                'ytdRateText': pct_text(x['ytdRate'] / 100 if x['ytdRate'] and x['ytdRate'] > 1 else x['ytdRate']),
            } for x in rate_low
        ]
    }


def main() -> None:
    PROCESSED.mkdir(parents=True, exist_ok=True)
    deposits = parse_deposit()
    loans = dedup_records(parse_loan(), key_fields=['bank', 'metricGroup', 'observedAt'])
    rates = parse_rates()
    summary = build_summary(deposits, loans, rates)
    quality = build_quality_report(deposits, loans, rates)
    dashboard = build_dashboard_payload(summary)

    write_json('deposit_benchmark.json', deposits)
    write_json('loan_benchmark.json', loans)
    write_json('loan_rate.json', rates)
    write_json('summary.json', summary)
    write_json('quality_report.json', quality)
    write_json('dashboard.json', dashboard)

    write_csv('deposit_benchmark.csv', deposits)
    write_csv('loan_benchmark.csv', loans)
    write_csv('loan_rate.csv', rates)

    print(f'wrote deposits={len(deposits)} loans={len(loans)} rates={len(rates)}')


if __name__ == '__main__':
    main()
